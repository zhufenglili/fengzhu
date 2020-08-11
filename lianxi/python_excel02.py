from openpyxl import workbook, load_workbook, Workbook


class Cc:

    def ww(self):
        wb = Workbook()
        # 获取当前活跃的页签
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = '体重'
        # 写入身高
        self.heiht = [190, 180, 170]
        wight = [66, 79, 30]
        for i in range(len(self.heiht)):
            #            行从第二行开始，每行加I
            d = ws.cell(row=2 + i, column=1, value=self.heiht[i])
            e = ws.cell(row=2 + i, column=2, value=wight[i])

        # 保存
        wb.save("sample.xlsx")
    def heiht_wight(self):
        ld = load_workbook("sample.xlsx") #一般用于读取
        sheet = ld.active#通过sheet 找到页签
        sheet ['c1'] = '备注'
        for i in range(len(self.heiht)):
         heiht = sheet.cell(row=2 + i, column=1).value #.value拿到VALUE
        weiht = sheet.cell(row=2 + i, column=1).value  # .value拿到VALUE
        #获取身高对应体重
        heiht_w = (heiht - 70)*0.6
        if weiht == heiht_w:
            print("这是健康的体重",heiht_w)
            sheet.cell(row=2+i,column=3).value="健康体重"
        ld.save("sample1.xlsx")#一定要保存

#调用时先实例化
aa = Cc()
aa.ww()
aa.heiht_wight()
