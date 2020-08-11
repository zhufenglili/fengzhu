from openpyxl import Workbook
wb = Workbook()

# 获取当前活跃的页签
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = "身高"
ws['B1'] = '体重'
#写入身高
heiht = [190,180,170]
wight = [66,79,30]
for i in range(len(heiht)):
    #            行从第二行开始，每行加I
    d = ws.cell(row=2+i, column=1, value=heiht[i])
    e = ws.cell(row=2 + i, column=2, value=wight[i])

#保存
wb.save("sample.xlsx")